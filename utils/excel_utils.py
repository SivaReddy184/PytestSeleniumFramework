import openpyxl
from openpyxl.styles import PatternFill


class ExcelUtils:

    @staticmethod
    def get_row_count(file_path, sheet_name):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.max_row

    @staticmethod
    def get_col_count(file_path, sheet_name):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.max_column

    @staticmethod
    def read_data(file_path, sheet_name, row_num, col_num):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row_num, column=col_num).value

    @staticmethod
    def write_data(file_path, sheet_name, row_num, col_num, data):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        sheet.cell(row=row_num, column=col_num).value = data
        workbook.save(file_path)


    @staticmethod
    def fill_green_color(file_path, sheet_name, row_num, col_num):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        green_fill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
        sheet.cell(row_num,col_num).fill = green_fill
        workbook.save(file_path)

    @staticmethod
    def get_data_from_excel(file, sheet_name):
        data_list = []
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        for row in range(2, sheet.max_row + 1):
            user = sheet.cell(row, 1).value
            pwd = sheet.cell(row, 2).value
            exp = sheet.cell(row, 3).value
            data_list.append((user, pwd, exp))
        return data_list

    # def get_data_from_excel(file_path, sheet_name):
    #     workbook = openpyxl.load_workbook(file_path)
    #     sheet = workbook[sheet_name]
    #     total_rows = sheet.max_row
    #     total_cols = sheet.max_column
    #     print(total_rows)
    #     print(total_cols)
    #
    #     for i in range(2, total_rows+1):
    #         for j in range(1, total_cols+1):
    #             print(sheet.cell(row=i, column=j).value)